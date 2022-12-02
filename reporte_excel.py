import openpyxl as xl
from openpyxl import chart, styles
import pandas as pd

reporte = pd.read_csv('weekly_ing.csv')
r = reporte.to_excel('reporte.xlsx', sheet_name='Reporte_de_pedidos_semanales',index=False)
wb = xl.load_workbook('reporte.xlsx')

title_font = xl.styles.Font(name='Arial', size=40, bold=True)
# subtitle color blue
subtitle_font = xl.styles.Font(name='Arial', size=18, underline='single', color='3366CC')
title_fill = xl.styles.PatternFill(fill_type='solid', start_color='73B6EE', end_color='73B6EE')
subtitle_fill = xl.styles.PatternFill(fill_type='solid', start_color='FFDB58', end_color='FFDB58')


ej_sheet = wb.create_sheet('Ejemplos_de_ingredientes',0)
ej_sheet.merge_cells('A1:J1')
ej_sheet['A1'] = 'Ejemplos de ingredientes'
ej_sheet['A1'].font = title_font
ej_sheet['A1'].fill = title_fill
ej_sheet['A1'].alignment = styles.Alignment(horizontal='center', vertical='center')

# Load the image
images = ['Genoa Salami.png',
    'Kalamata Olives.png',
    'Onions.png']

positions = ['B3', 'B29', 'B55']
for i in range(len(images)):
    img = xl.drawing.image.Image(images[i])
    ej_sheet.add_image(img, positions[i])

ult_pedido = wb.create_sheet('Pedido_siguiente_semana',0)
ult_pedido.merge_cells('A1:L1')
ult_pedido['A1'] = 'Pedido de la siguiente semana'
ult_pedido['A1'].font = title_font
ult_pedido['A1'].fill = title_fill
# make the column width bigger
ult_pedido.column_dimensions['A'].width = 27
ult_pedido.column_dimensions['B'].width = 18
ult_pedido['A1'].alignment = styles.Alignment(horizontal='center', vertical='center')

ult_pedido['A2'] = 'Ingredientes'
ult_pedido['A2'].font = styles.Font(name='Arial', size=18, underline='single', bold=True)
ult_pedido['B2'] = 'Cantidad'
ult_pedido['B2'].font = styles.Font(name='Arial', size=18, underline='single', bold=True)
ult = reporte[reporte['week'] == 104]
ult = ult.drop(['week'], axis=1)
cantidades = {}

for i in ult:
    cantidades[i] = ult[i].values[0]
cantidades = sorted(cantidades.items(), key=lambda x: x[1], reverse=True)
# create table with ingredients and quantities
for i in range(len(cantidades)):
    ult_pedido.cell(row=i+3, column=1, value=cantidades[i][0])
    ult_pedido.cell(row=i+3, column=2, value=cantidades[i][1])
    ult_pedido.cell(row=i+3, column=2).alignment = styles.Alignment(horizontal='center', vertical='center')

# create a chart
chart1 = chart.BarChart()
chart1.type = "bar"
chart1.style = 10
chart1.title = "Cantidad de ingredientes en el utimo pedido"
chart1.y_axis.title = 'Cantidad'
chart1.x_axis.title = 'Ingredientes'
data = chart.Reference(ult_pedido, min_col=2, min_row=2, max_col=2, max_row=67)
cats = chart.Reference(ult_pedido, min_col=1, min_row=3, max_row=67)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.height = 38
chart1.width = 20
chart1.legend = None
chart1.varyColors = True
ult_pedido.add_chart(chart1, "C2")



title_page = wb.create_sheet('Portada',0)
title_page.merge_cells('A1:J1')
title_page['A1'] = "\U0001F355 Maven Pizza Report \U0001F355"
title_page['A1'].font = title_font
title_page['A1'].fill = title_fill
title_page['A1'].alignment = styles.Alignment(horizontal='center', vertical='center')


link = "reporte.xlsx#Pedido_siguiente_semana!A1"
title_page.cell(row=3, column=3).hyperlink = (link)
title_page.merge_cells('C3:H3')
title_page['C3'] = 'Ir al pedido de la semana que viene'
title_page['C3'].font = subtitle_font
title_page['C3'].alignment = styles.Alignment(horizontal='center', vertical='center')
title_page['C3'].fill = subtitle_fill

link = "reporte.xlsx#Ejemplos_de_ingredientes!A1"
title_page.cell(row=5, column=3).hyperlink = (link)
title_page.merge_cells('C5:H5')
title_page['C5'] = 'Ir a ejemplos de stock de ingredientes'
title_page['C5'].font = subtitle_font
title_page['C5'].alignment = styles.Alignment(horizontal='center', vertical='center')
title_page['C5'].fill = subtitle_fill

link = "reporte.xlsx#Reporte_de_pedidos_semanales!A1"
title_page.cell(row=7, column=3).hyperlink = (link)
title_page.merge_cells('C7:H7')
title_page['C7'] = 'Ir a todos los pedidos'
title_page['C7'].font = subtitle_font
title_page['C7'].alignment = styles.Alignment(horizontal='center', vertical='center')
title_page['C7'].fill = subtitle_fill

img = xl.drawing.image.Image('pizza.jpeg')
img.width *= 10/14 
img.height *= 10/14 

title_page.add_image(img, 'A9')





wb.save('reporte.xlsx')